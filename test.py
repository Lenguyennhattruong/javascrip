import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from main import tinh_tong

def chon_file():
    file_path = filedialog.askopenfilename(
        title="Chọn file Excel test",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    entry_path.delete(0, tk.END)
    entry_path.insert(0, file_path)

def chay_kiem_thu():
    file_path = entry_path.get()
    if not file_path:
        messagebox.showerror("Lỗi", "Vui lòng chọn file Excel test trước!")
        return

    try:
        wb = load_workbook(file_path)
        sheet = wb.active

        pass_count = 0
        fail_count = 0

        for row in range(2, sheet.max_row + 1):
            a = sheet.cell(row, 1).value
            b = sheet.cell(row, 2).value
            expected = sheet.cell(row, 3).value

            result = tinh_tong(a, b)
            if result == expected:
                sheet.cell(row, 4).value = "✅ Pass"
                pass_count += 1
            else:
                sheet.cell(row, 4).value = f"❌ Fail (got {result})"
                fail_count += 1

        wb.save(file_path)

        messagebox.showinfo(
            "Hoàn tất",
            f"Đã kiểm thử xong!\n✅ Pass: {pass_count}\n❌ Fail: {fail_count}"
        )
    except Exception as e:
        messagebox.showerror("Lỗi khi chạy test", str(e))

# === Giao diện chính ===
root = tk.Tk()
root.title("🧪 Chương trình kiểm thử Excel")
root.geometry("500x250")
root.resizable(False, False)

# Nhãn & ô nhập
label = tk.Label(root, text="Chọn file Excel test:", font=("Arial", 12))
label.pack(pady=10)

frame = tk.Frame(root)
frame.pack()

entry_path = tk.Entry(frame, width=50, font=("Arial", 10))
entry_path.pack(side=tk.LEFT, padx=5)

btn_chon = tk.Button(frame, text="📂 Chọn file", command=chon_file)
btn_chon.pack(side=tk.LEFT, padx=5)

# Nút chạy kiểm thử
btn_run = tk.Button(
    root, text="▶️ Chạy kiểm thử",
    bg="#4CAF50", fg="white",
    font=("Arial", 12, "bold"),
    command=chay_kiem_thu
)
btn_run.pack(pady=20, ipadx=10, ipady=5)

# Nhãn thông tin
footer = tk.Label(root, text="© Kiểm thử tự động từ Excel - Python tkinter", fg="gray")
footer.pack(side=tk.BOTTOM, pady=5)

root.mainloop()
